import json
import warnings
import pandas as pd
import openpyxl
from datetime import datetime
from pathlib import Path
from src.mlac_transformer.logger import write_log


class ExcelToJson:
    """
    Transforms an Excel file to a JSON representation.
    """

    def __init__(self, excel_path: str) -> None:
        today = datetime.today()
        self.excel_path = excel_path
        self.output_path = (
            Path("output/extraction")
            / today.strftime("%Y")
            / today.strftime("%m")
            / today.strftime("%d")
        )

    def run(self) -> str:
        try:
            with warnings.catch_warnings():
                warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")
                wb = openpyxl.load_workbook(self.excel_path)
                all_sheets = pd.read_excel(self.excel_path, sheet_name=None)
        except Exception as e:
            write_log("error", f"Failed to read Excel file '{self.excel_path}': {e}")
            raise

        raw_data = {}

        for sheet_name, df in all_sheets.items():
            try:
                ws = wb[sheet_name]
                df, group_id_map = self._resolve_merged_cells(ws, df)

                # Ignore 'Unnamed' columns
                df = df.loc[:, ~df.columns.astype(str).str.startswith('Unnamed')]
                df['__group_id__'] = [group_id_map[i] for i in range(len(df))]

                # Collapse each group into one record, joining unique non-empty values with 'comma-space'
                def join_unique(series):
                    seen = []
                    for v in series:
                        s = str(v).strip()
                        if s and s != 'nan' and s not in seen:
                            seen.append(s)
                    return ', '.join(seen)

                raw_data[sheet_name] = (
                    df.groupby('__group_id__', sort=True)
                    .agg(join_unique)
                    .reset_index(drop=True)
                    .to_dict(orient='records')
                )
            except Exception as e:
                write_log("error", f"Failed to process sheet '{sheet_name}' in '{self.excel_path}': {e}")
                raise

        try:
            # Build output path: output/extraction/YYYY/MM/DD/<excel_filename>.json
            self.output_path.mkdir(parents=True, exist_ok=True)
            raw_file = self.output_path / (Path(self.excel_path).stem + ".json")
            raw_file.write_text(
                json.dumps({"workbook": raw_data}, indent=2),
                encoding="utf-8"
            )
        except Exception as e:
            write_log("error", f"Failed to write JSON output for '{self.excel_path}': {e}")
            raise

        return str(raw_file)

    def _resolve_merged_cells(self, ws, df):
        """
        Fill merged cell values in df and return a group_id_map for collapsing rows.
        """
        group_id_map = {i: i for i in range(len(df))}
        for merged_range in ws.merged_cells.ranges:
            top_left = ws.cell(merged_range.min_row, merged_range.min_col).value
            first_df_row = None
            for row in range(merged_range.min_row, merged_range.max_row + 1):
                df_row = row - 2
                if 0 <= df_row < len(df):
                    for col in range(merged_range.min_col, merged_range.max_col + 1):
                        col_idx = col - 1
                        if col_idx < len(df.columns):
                            df.iat[df_row, col_idx] = top_left
                    if merged_range.max_row > merged_range.min_row:
                        if first_df_row is None:
                            first_df_row = df_row
                        else:
                            group_id_map[df_row] = group_id_map[first_df_row]
        return df, group_id_map
