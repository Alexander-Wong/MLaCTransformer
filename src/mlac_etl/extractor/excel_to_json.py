import json
import re
import warnings
import openpyxl
from openpyxl.cell.cell import MergedCell
from xlcalculator import ModelCompiler, Evaluator
import mlac_etl.extractor.xl_custom_functions  # noqa: F401 — registers custom xl functions at import time
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from mlac_etl.logger import write_log

_RE_AUTHOR_TIMESTAMP = re.compile(r".+\s{2,}\(\d{4}-\d{2}-\d{2}")
_RE_TRAILING_PARENS = re.compile(r"\s*\([^)]*\)\s*$")
_RE_KEY_SUFFIX = re.compile(r"\(Key\)\s*$")


class ExcelToJson:
    """
    Transforms an Excel file to a JSON representation.

    Every cell value is represented as an object:
        { "value": "<cell text>" }

    If the cell has a comment, one additional key is added:
        { "value": "...", "annotation": "<comment text>" }

    Formula cells are evaluated via xlcalculator (lazy-initialized only when
    formulas are detected). Relational data is resolved through the Universal
    Matrix Mapping engine using [MAP] columns and (Key) identifiers.
    """

    def __init__(self, excel_path: str) -> None:
        """Store the input path and resolve the timestamped output directory."""
        self.today = datetime.today()
        self.excel_path = excel_path
        self._evaluator = None
        self.output_path = (
            Path("output/extraction")
            / self.today.strftime("%Y")
            / self.today.strftime("%m")
            / self.today.strftime("%d")
        )

    def run(self) -> str:
        """Load the workbook, process every sheet, and write the output JSON. Returns the output file path."""
        try:
            with warnings.catch_warnings():
                warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")
                wb = openpyxl.load_workbook(self.excel_path)
        except Exception as e:
            write_log("error", f"Failed to read Excel file '{self.excel_path}': {e}")
            raise

        if self._workbook_has_formulas(wb):
            try:
                compiler = ModelCompiler()
                xc_model = compiler.read_and_parse_archive(self.excel_path)
                self._evaluator = Evaluator(xc_model)
            except Exception as e:
                write_log("warning", f"xlcalculator failed to initialize for '{self.excel_path}': {e}. Formula cells will return raw formula strings.")

        raw_data = {}

        for sheet_name in wb.sheetnames:
            try:
                ws = wb[sheet_name]
                rows, group_id_map = self._build_sheet_rows(ws, sheet_name)
                raw_data[sheet_name] = self._collapse_groups(rows, group_id_map)
            except Exception as e:
                write_log("error", f"Failed to process sheet '{sheet_name}' in '{self.excel_path}': {e}")
                raise

        try:
            raw_data = self._apply_universal_matrix_mapping(raw_data)

            raw_data_copy = {k: v for k, v in raw_data.items()}
            raw_data["__GLOBAL_WORKBOOK__"] = [{"__FULL_WORKBOOK__": {"value": json.dumps(raw_data_copy)}}]

            write_log("info", "Global Workbook successfully injected.")
        except Exception as e:
            write_log("error", f"Failed to run Matrix Mapping or Global Workbook injection: {e}")
            raise

        try:
            self.output_path.mkdir(parents=True, exist_ok=True)
            raw_file = self.output_path / (Path(self.excel_path).stem + "-" + self.today.strftime("%H-%M-%S") + ".json")
            raw_file.write_text(
                json.dumps({"workbook": raw_data}, indent=2),
                encoding="utf-8"
            )
        except Exception as e:
            write_log("error", f"Failed to write JSON output for '{self.excel_path}': {e}")
            raise

        return str(raw_file)

    # =========================================================================
    # UNIVERSAL MATRIX MAPPING
    # =========================================================================

    def _val(self, cell_obj) -> str:
        """Extract the string value from a cell object dict or a raw value."""
        if isinstance(cell_obj, dict):
            return str(cell_obj.get("value", "")).strip()
        return str(cell_obj or "").strip()

    def _find_primary_key(self, row: dict) -> str:
        """Return the value of the first column whose name contains '(Key)', or None if absent."""
        for col_name, cell_obj in row.items():
            if "(Key)" in str(col_name):
                return self._val(cell_obj)
        return None

    def _apply_universal_matrix_mapping(self, raw_data: dict) -> dict:
        """Resolve cross-sheet relations defined by [MAP] columns and (Key) identifiers.

        Pre-scan builds map_cols (which sheets/columns have [MAP]) and key_index
        ({key_value → row_ref}) so each pass only touches relevant data.
        PASS 1 builds the relations index restricted to MAP sheets and MAP columns.
        PASS 2 injects via key_index in O(relations) instead of O(S × R).
        """
        write_log("info", "Executing Universal Matrix Mapping...")

        # Pre-scan: O(S × C) — headers only, no row data
        map_cols  = {}  # sheet_name → [col_names that start with [MAP]]
        key_index = {}  # key_value  → direct row reference in raw_data

        for sheet_name, rows in raw_data.items():
            if not rows:
                continue
            first_keys = list(rows[0].keys())
            col_names = [k for k in first_keys if str(k).upper().startswith("[MAP]")]
            if col_names:
                map_cols[sheet_name] = col_names
            key_col = next((k for k in first_keys if "(Key)" in str(k)), None)
            if key_col:
                for row in rows:
                    key_val = self._val(row.get(key_col, {}))
                    if key_val:
                        key_index[key_val] = row

        if not map_cols:
            write_log("info", "No [MAP] columns found. Returning flat JSON.")
            return raw_data

        relations = defaultdict(lambda: defaultdict(list))
        maps_found = 0

        # PASS 1: only MAP sheets, only MAP columns — O(S_map × R × C_map)
        for sheet_name, col_names in map_cols.items():
            for row in raw_data[sheet_name]:
                for col_name in col_names:
                    if col_name not in row:
                        continue
                    target_id = col_name[5:].strip()
                    val = self._val(row[col_name]).strip().upper()
                    if val and val not in ["0", "FALSE", "NO", "N"]:
                        row_copy = row.copy()
                        row_copy["__MAP_VALUE__"] = val
                        relations[target_id][sheet_name].append(row_copy)
                        maps_found += 1

        if maps_found == 0:
            write_log("info", "No [MAP] relations found. Returning flat JSON.")
            return raw_data

        # PASS 2: inject via key_index — O(relations) instead of O(S × R)
        for target_id, sources in relations.items():
            if target_id not in key_index:
                continue
            row = key_index[target_id]
            for source_sheet, mapped_rows in sources.items():
                safe_sheet_name = re.sub(r'[^a-zA-Z0-9]', '_', source_sheet)
                row[f"__MAPPED_{safe_sheet_name}"] = mapped_rows

        write_log("info", f"Matrix mapping complete. {maps_found} relations injected.")
        return raw_data

    # =========================================================================
    # SHEET PROCESSING
    # =========================================================================

    def _build_sheet_rows(self, ws, sheet_name: str = "") -> tuple:
        """Read a worksheet into row dicts and a group_id_map for merged-row grouping."""
        if ws.max_row < 2:
            return [], {}

        headers = [
            self._normalize_header(str(cell.value)) if cell.value is not None else None
            for cell in ws[1]
        ]

        n_rows = ws.max_row - 1
        rows = [{} for _ in range(n_rows)]

        merged_top_lefts = {(r.min_row, r.min_col) for r in ws.merged_cells.ranges}

        for df_row, row_cells in enumerate(ws.iter_rows(min_row=2)):
            excel_row = df_row + 2
            for col_idx, cell in enumerate(row_cells, start=1):
                header = headers[col_idx - 1]
                if not header or header.startswith("Unnamed"):
                    continue
                if not isinstance(cell, MergedCell) and (excel_row, col_idx) not in merged_top_lefts:
                    rows[df_row][header] = self._build_cell_object(cell, sheet_name)

        group_id_map = {i: i for i in range(n_rows)}
        for merged_range in ws.merged_cells.ranges:
            top_left = ws.cell(merged_range.min_row, merged_range.min_col)
            cell_obj = self._build_cell_object(top_left, sheet_name)
            first_df_row = None
            for excel_row in range(merged_range.min_row, merged_range.max_row + 1):
                df_row = excel_row - 2
                if 0 <= df_row < n_rows:
                    for col in range(merged_range.min_col, merged_range.max_col + 1):
                        header = headers[col - 1]
                        if not header or header.startswith("Unnamed"):
                            continue
                        rows[df_row][header] = cell_obj
                    if merged_range.max_row > merged_range.min_row:
                        if first_df_row is None:
                            first_df_row = df_row
                        else:
                            group_id_map[df_row] = group_id_map[first_df_row]

        return rows, group_id_map

    def _build_cell_object(self, cell, sheet_name: str = "") -> dict:
        """Return `{"value": ...}`, adding `"annotation"` only when a comment is present."""
        obj = {"value": self._clean_value(self._resolve_cell_value(cell, sheet_name))}
        if cell.comment and cell.comment.text:
            obj["annotation"] = self._parse_comment(cell.comment.text, cell.comment.author)
        return obj

    def _resolve_cell_value(self, cell, sheet_name: str):
        """Return the evaluated result for formula cells; return raw value otherwise."""
        value = cell.value
        if self._evaluator is None or not (isinstance(value, str) and value.startswith("=")):
            return value
        try:
            quote = "'" in sheet_name or " " in sheet_name
            ref = f"'{sheet_name}'!{cell.coordinate}" if quote else f"{sheet_name}!{cell.coordinate}"
            return self._evaluator.evaluate(ref)
        except Exception as e:
            write_log("warning", f"Could not evaluate formula '{value}' at {sheet_name}!{cell.coordinate}: {e}")
            return value

    def _collapse_groups(self, rows: list, group_id_map: dict) -> list:
        """Merge rows that share the same group_id (vertical merge group) into a single record."""
        groups = defaultdict(list)
        for i, row in enumerate(rows):
            groups[group_id_map[i]].append(row)

        result = []
        for group_id in sorted(groups.keys()):
            group_rows = groups[group_id]
            seen_keys = dict.fromkeys(k for r in group_rows for k in r)
            merged = {
                key: self._merge_cell_objects([r[key] for r in group_rows if key in r])
                for key in seen_keys
            }
            result.append(merged)
        return result

    def _merge_cell_objects(self, objects: list) -> dict:
        """Combine cell objects from the same group: unique non-empty values joined, first annotation wins."""
        seen_values = {}
        annotation = None

        for obj in objects:
            v = obj.get("value", "").strip()
            if v and v != "nan" and v not in seen_values:
                seen_values[v] = None
            if annotation is None and obj.get("annotation"):
                annotation = obj["annotation"]

        result = {"value": ", ".join(seen_values)}
        if annotation:
            result["annotation"] = annotation
        return result

    # =========================================================================
    # STATIC HELPERS
    # =========================================================================

    @staticmethod
    def _normalize_header(header: str | None) -> str | None:
        """
        Canonicalize Excel column headers once at extraction time.

        Preserves the semantic ``(Key)`` suffix (primary row identifier).
        Strips any other trailing parenthetical suffix — e.g. ``(Drop)``,
        ``(Reference)``, ``(Cutline)``, hints on ``[MAP]`` slugs — so YAML/jq
        can reference stable names without Excel metadata noise.
        """
        if header is None:
            return None
        h = str(header).strip()
        if not h or h.startswith("Unnamed"):
            return h

        if h.upper().startswith("[MAP]"):
            bracket = h.find("]")
            if bracket == -1:
                return h
            prefix = h[: bracket + 1]
            slug = h[bracket + 1 :].strip()
            while True:
                cleaned = _RE_TRAILING_PARENS.sub("", slug).strip()
                if cleaned == slug:
                    break
                slug = cleaned
            return f"{prefix} {slug}".strip() if slug else prefix

        if _RE_KEY_SUFFIX.search(h):
            return h

        while True:
            cleaned = _RE_TRAILING_PARENS.sub("", h).strip()
            if cleaned == h:
                break
            h = cleaned
        return h

    @staticmethod
    def _workbook_has_formulas(wb) -> bool:
        """Return True as soon as a formula cell is found; avoids full scan when unnecessary."""
        for sheet_name in wb.sheetnames:
            for row in wb[sheet_name].iter_rows(min_row=2):
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        return True
        return False

    @staticmethod
    def _parse_comment(text: str, author: str = "") -> str:
        """Strip Excel metadata from a comment and return only the body text."""
        text = (text or "").strip()

        if "======" in text:
            return "\n".join(
                line for line in text.split("\n")
                if not line.startswith("======")
                and not line.startswith("ID#")
                and not _RE_AUTHOR_TIMESTAMP.match(line)
            ).strip()

        if author:
            prefix = f"{author}:\n"
            if text.startswith(prefix):
                text = text[len(prefix):]
        return text.strip()

    @staticmethod
    def _clean_value(value) -> str:
        """Convert a cell value to a clean string."""
        if value is None:
            return ""
        return str(value).strip()
